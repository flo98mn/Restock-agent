#!/usr/bin/env python3
"""
stock_analyst.py — pasul de INGESTĂ din pipeline-ul Restock-agent.

Determinist, fără LLM, fără rețea, fără DB. Citește cel mai recent
"Raport stoc 3.0_*.xlsx" din REPORT_DIR, parsează sheet-urile Stoc_viitor și
Restock_tehnic DUPĂ NUMELE din header (nu după index fix), face join pe Nickname
și scrie data/cache/restock_dataset_<report_date>.json.

NU calculează nicio cantitate de comandat. NU atinge fișierul sursă.
Vezi context/business-context.md pentru detaliile de business.
"""

import argparse
import glob
import json
import os
import re
import sys
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "openpyxl lipsește. Instalează: python3 -m pip install openpyxl\n"
    )
    sys.exit(1)

REPORT_GLOB = "Raport stoc 3.0_*.xlsx"
DATE_RE = re.compile(r"(\d{4}_\d{2}_\d{2})")

SHEET_STOC = "Stoc_viitor"
SHEET_TEHNIC = "Restock_tehnic"

REQ_STOC = ["Nickname", "Current Stock", "PCS Sold in Last 30 Days"]
REQ_TEHNIC = ["NicknameID", "Nickname", "Furnizor", "Baterie", "Multiplu"]
OPT_TEHNIC = ["Greutate_kg", "Volum_cm3"]


def die(msg):
    sys.stderr.write("EROARE: " + msg + "\n")
    sys.exit(1)


def to_num(v, default=0):
    """Coerce la număr; None / non-numeric -> default."""
    if v is None:
        return default
    if isinstance(v, bool):  # bool e subclasă de int; tratează ca int
        return int(v)
    if isinstance(v, (int, float)):
        return v
    try:
        s = str(v).strip().replace(",", ".")
        if s == "":
            return default
        f = float(s)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return default


def to_int_or_none(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    n = to_num(v, None)
    if n is None:
        return None
    return int(n)


def header_map(ws):
    """Rândul 1 -> {nume_header: index}. Strip pe text."""
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hmap = {}
    for i, h in enumerate(hdr):
        if h is None:
            continue
        name = str(h).strip()
        if name and name not in hmap:
            hmap[name] = i
    return hmap, hdr


def find_report(report_dir, explicit):
    if explicit:
        if not os.path.isfile(explicit):
            die("fișierul --report nu există: " + explicit)
        return explicit
    pattern = os.path.join(report_dir, REPORT_GLOB)
    candidates = glob.glob(pattern)
    if not candidates:
        die("niciun fișier '" + REPORT_GLOB + "' în " + report_dir)
    dated = []
    for c in candidates:
        m = DATE_RE.search(os.path.basename(c))
        if m:
            dated.append((m.group(1), c))
    if not dated:
        die("niciun fișier cu dată (YYYY_MM_DD) în nume în " + report_dir)
    dated.sort(key=lambda t: t[0])  # cel mai recent la final
    return dated[-1][1]


def parse_stoc_viitor(ws):
    """Returnează (lista produse parțiale per nickname, order_columns, warnings)."""
    hmap, _ = header_map(ws)
    missing = [c for c in REQ_STOC if c not in hmap]
    if missing:
        die(SHEET_STOC + ": coloane obligatorii lipsă: " + ", ".join(missing))

    order_cols = [name for name in hmap if name.startswith("Order:")]
    # ordonează după indexul coloanei pentru stabilitate
    order_cols.sort(key=lambda n: hmap[n])
    order_idx = [hmap[n] for n in order_cols]

    i_nick = hmap["Nickname"]
    i_stock = hmap["Current Stock"]
    i_sold = hmap["PCS Sold in Last 30 Days"]

    products = []
    warnings = []
    seen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if i_nick >= len(row):
            continue
        nick = row[i_nick]
        if nick is None or str(nick).strip() == "":
            continue
        nick = str(nick).strip()

        incoming = 0
        for idx in order_idx:
            if idx < len(row):
                incoming += to_num(row[idx], 0)

        rec = {
            "nickname": nick,
            "current_stock": to_num(row[i_stock] if i_stock < len(row) else None, 0),
            "pcs_sold_30d": to_num(row[i_sold] if i_sold < len(row) else None, 0),
            "incoming": incoming,
        }
        if nick in seen:
            warnings.append("Nickname duplicat în " + SHEET_STOC + ": " + nick)
        seen[nick] = rec
        products.append(rec)

    return products, order_cols, warnings


def parse_restock_tehnic(ws):
    """Returnează ({nickname: tehnic_rec}, nr_rows, warnings)."""
    hmap, _ = header_map(ws)
    missing = [c for c in REQ_TEHNIC if c not in hmap]
    if missing:
        die(SHEET_TEHNIC + ": coloane obligatorii lipsă: " + ", ".join(missing))

    idx = {c: hmap[c] for c in REQ_TEHNIC}
    for c in OPT_TEHNIC:
        idx[c] = hmap.get(c)

    by_nick = {}
    warnings = []
    n_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        i_nick = idx["Nickname"]
        if i_nick >= len(row):
            continue
        nick = row[i_nick]
        if nick is None or str(nick).strip() == "":
            continue
        nick = str(nick).strip()
        n_rows += 1

        def cell(col):
            i = idx[col]
            if i is None or i >= len(row):
                return None
            return row[i]

        rec = {
            "nickname_id": to_int_or_none(cell("NicknameID")),
            "supplier": (str(cell("Furnizor")).strip()
                         if cell("Furnizor") is not None else None),
            "battery": int(to_num(cell("Baterie"), 0)),
            "multiple": int(to_num(cell("Multiplu"), 0)),
            "weight_kg": cell("Greutate_kg"),
            "volume_cm3": cell("Volum_cm3"),
        }
        if rec["weight_kg"] is not None:
            rec["weight_kg"] = to_num(rec["weight_kg"], None)
        if rec["volume_cm3"] is not None:
            rec["volume_cm3"] = to_num(rec["volume_cm3"], None)

        if nick in by_nick:
            warnings.append("Nickname duplicat în " + SHEET_TEHNIC + ": " + nick)
        by_nick[nick] = rec

    return by_nick, n_rows, warnings


def main():
    ap = argparse.ArgumentParser(description="Restock-agent: ingestă Stoc_viitor + Restock_tehnic.")
    ap.add_argument("--report", help="cale către un xlsx anume (override auto-detect)")
    args = ap.parse_args()

    report_dir = os.environ.get("REPORT_DIR", "./data/reports")
    report_file = find_report(report_dir, args.report)

    m = DATE_RE.search(os.path.basename(report_file))
    report_date = m.group(1) if m else "unknown"

    wb = openpyxl.load_workbook(report_file, read_only=True, data_only=True)
    for sn in (SHEET_STOC, SHEET_TEHNIC):
        if sn not in wb.sheetnames:
            die("sheet lipsă: " + sn)

    stoc_products, order_cols, w1 = parse_stoc_viitor(wb[SHEET_STOC])
    tehnic_by_nick, tehnic_rows, w2 = parse_restock_tehnic(wb[SHEET_TEHNIC])
    wb.close()

    warnings = list(w1) + list(w2)

    # JOIN pe Nickname
    products = []
    unmatched_in_stoc = []  # în Stoc_viitor, lipsă în Restock_tehnic
    stoc_nicks = set()
    for p in stoc_products:
        nick = p["nickname"]
        stoc_nicks.add(nick)
        t = tehnic_by_nick.get(nick)
        if t is None:
            unmatched_in_stoc.append(nick)
            continue  # NU inventa valori; exclus din products
        products.append({
            "nickname_id": t["nickname_id"],
            "nickname": nick,
            "current_stock": p["current_stock"],
            "pcs_sold_30d": p["pcs_sold_30d"],
            "incoming": p["incoming"],
            "supplier": t["supplier"],
            "battery": t["battery"],
            "multiple": t["multiple"],
            "weight_kg": t["weight_kg"],
            "volume_cm3": t["volume_cm3"],
        })

    # în Restock_tehnic, lipsă în Stoc_viitor
    unmatched_in_tehnic = sorted(n for n in tehnic_by_nick if n not in stoc_nicks)
    unmatched_in_stoc = sorted(unmatched_in_stoc)

    products.sort(key=lambda r: (r["nickname_id"] is None, r["nickname_id"] or 0, r["nickname"]))

    meta = {
        "report_file": os.path.basename(report_file),
        "report_date": report_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "stoc_viitor_rows": len(stoc_products),
        "restock_tehnic_rows": tehnic_rows,
        "matched": len(products),
        "unmatched_in_stoc_viitor": unmatched_in_stoc,
        "unmatched_in_restock_tehnic": unmatched_in_tehnic,
        "order_columns_detected": order_cols,
        "data_warnings": warnings,
    }

    out_dir = os.path.join("data", "cache")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "restock_dataset_" + report_date + ".json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "products": products}, f, ensure_ascii=False, indent=2)

    # Sumar stdout
    print("=== stock_analyst ===")
    print("report_file :", meta["report_file"])
    print("report_date :", meta["report_date"])
    print("stoc_viitor_rows :", meta["stoc_viitor_rows"])
    print("restock_tehnic_rows :", meta["restock_tehnic_rows"])
    print("matched :", meta["matched"])
    print("unmatched_in_stoc_viitor :", len(unmatched_in_stoc), unmatched_in_stoc[:10])
    print("unmatched_in_restock_tehnic :", len(unmatched_in_tehnic), unmatched_in_tehnic[:10])
    print("order_columns_detected :", len(order_cols))
    for oc in order_cols:
        print("   -", oc)
    print("output :", out_path)
    print("--- exemple produse ---")
    for p in products[:3]:
        print("  ", json.dumps(p, ensure_ascii=False))


if __name__ == "__main__":
    main()
