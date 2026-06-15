#!/usr/bin/env python3
"""
report_writer.py — pasul de SCRIERE din pipeline-ul Restock-agent.

Determinist, fără LLM/rețea/DB. Copiază fișierul sursă "Raport stoc 3.0_<date>.xlsx"
și scrie DOAR cantitățile noastre (Comanda aer / Comanda tren) în sheet-ul Stoc_viitor,
păstrând toate celelalte sheet-uri și formulele intacte. Marchează celulele Kaka high-risk
și adaugă un sheet de justificare "Restock_calcul".

NU modifică fișierul sursă din data/reports/. NU recalculează formule (le lasă ca string;
Excel le recalculează la deschidere). Vezi context/business-context.md.
"""

import argparse
import glob
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
except ImportError:
    sys.stderr.write("openpyxl lipsește. Instalează: python -m pip install openpyxl\n")
    sys.exit(1)

DATE_RE = re.compile(r"(\d{4}_\d{2}_\d{2})")
PLAN_GLOB = os.path.join("data", "cache", "restock_plan_*.json")

SHEET_STOC = "Stoc_viitor"
SHEET_CALCUL = "Restock_calcul"

KAKA_FILL = PatternFill(fill_type="solid", fgColor="FFC000")
KAKA_COMMENT = "Comandă aer SEPARATĂ high-risk (Kaka)"

CALCUL_HEADERS = [
    "NicknameID", "Nickname", "Furnizor", "Ruta", "Battery_source", "Kaka_high_risk",
    "Stoc_curent", "Pe_drum", "Vandut_30z", "V_zilnic", "Multiplu",
    "Raw_aer", "Comanda_aer", "Raw_tren", "Comanda_tren",
]


def die(msg):
    sys.stderr.write("EROARE: " + msg + "\n")
    sys.exit(1)


def find_plan(explicit):
    if explicit:
        if not os.path.isfile(explicit):
            die("fișierul --plan nu există: " + explicit)
        return explicit
    candidates = glob.glob(PLAN_GLOB)
    if not candidates:
        die("niciun 'restock_plan_*.json' în data/cache/ (rulează întâi restock-planner)")
    dated = []
    for c in candidates:
        m = DATE_RE.search(os.path.basename(c))
        if m:
            dated.append((m.group(1), c))
    if not dated:
        die("niciun plan cu dată (YYYY_MM_DD) în nume în data/cache/")
    dated.sort(key=lambda t: t[0])
    return dated[-1][1]


def header_map(ws):
    """Rândul 1 -> {nume_header: index 0-based}."""
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hmap = {}
    for i, h in enumerate(hdr):
        if h is None:
            continue
        name = str(h).strip()
        if name and name not in hmap:
            hmap[name] = i
    return hmap


def main():
    ap = argparse.ArgumentParser(description="Restock-agent: scrie cantitățile în Stoc_viitor.")
    ap.add_argument("--plan", help="cale către un plan JSON anume (override auto-detect)")
    args = ap.parse_args()

    plan_path = find_plan(args.plan)
    with open(plan_path, "r", encoding="utf-8") as f:
        plan = json.load(f)

    report_date = plan.get("meta", {}).get("report_date")
    if not report_date:
        m = DATE_RE.search(os.path.basename(plan_path))
        report_date = m.group(1) if m else None
    if not report_date:
        die("nu pot determina report_date din plan")

    report_dir = os.environ.get("REPORT_DIR", "./data/reports")
    source_path = os.path.join(report_dir, "Raport stoc 3.0_" + report_date + ".xlsx")
    if not os.path.isfile(source_path):
        die("fișierul sursă lipsește (trebuie EXACT raportul planului): " + source_path)

    plan_products = plan.get("products", [])
    by_nick = {}
    for p in plan_products:
        by_nick.setdefault(p["nickname"], p)

    # Încarcă sursa PĂSTRÂND formulele (data_only=False, nu read_only)
    wb = openpyxl.load_workbook(source_path)
    if SHEET_STOC not in wb.sheetnames:
        die("sheet lipsă în sursă: " + SHEET_STOC)
    ws = wb[SHEET_STOC]

    hmap = header_map(ws)
    for col in ("Nickname", "Comanda aer", "Comanda tren"):
        if col not in hmap:
            die(SHEET_STOC + ": coloană lipsă: " + col)
    i_nick = hmap["Nickname"]
    c_aer = hmap["Comanda aer"] + 1   # openpyxl e 1-based pe celule
    c_tren = hmap["Comanda tren"] + 1

    warnings = []
    seen = set()
    n_aer = 0
    n_tren = 0
    n_kaka_marked = 0

    for row_cells in ws.iter_rows(min_row=2):
        nick_cell = row_cells[i_nick] if i_nick < len(row_cells) else None
        nick = nick_cell.value if nick_cell is not None else None
        if nick is None or str(nick).strip() == "":
            continue
        nick = str(nick).strip()
        r = row_cells[0].row

        if nick in seen:
            warnings.append("Nickname duplicat în " + SHEET_STOC + ": " + nick)
        seen.add(nick)

        p = by_nick.get(nick)
        if p is None:
            warnings.append("Nickname din Stoc_viitor fără plan: " + nick)
            continue

        qty_aer = p.get("qty_aer", 0) or 0
        qty_tren = p.get("qty_tren", 0) or 0

        cell_aer = ws.cell(row=r, column=c_aer)
        cell_tren = ws.cell(row=r, column=c_tren)

        cell_aer.value = qty_aer if qty_aer > 0 else None
        cell_tren.value = qty_tren if qty_tren > 0 else None

        if p.get("route") == "kaka" and p.get("kaka_high_risk") and qty_aer > 0:
            cell_aer.fill = KAKA_FILL
            cell_aer.comment = Comment(KAKA_COMMENT, "restock-agent")
            n_kaka_marked += 1

        if qty_aer > 0:
            n_aer += 1
        if qty_tren > 0:
            n_tren += 1

    # Sheet de justificare Restock_calcul (ultimul sheet)
    if SHEET_CALCUL in wb.sheetnames:
        del wb[SHEET_CALCUL]
    wsc = wb.create_sheet(SHEET_CALCUL)
    wsc.append(CALCUL_HEADERS)
    for c in wsc[1]:
        c.font = Font(bold=True)

    ordered = sorted(
        plan_products,
        key=lambda x: (x.get("nickname_id") is None, x.get("nickname_id") or 0, x.get("nickname", "")),
    )
    for p in ordered:
        wsc.append([
            p.get("nickname_id"),
            p.get("nickname"),
            p.get("supplier"),
            p.get("route"),
            p.get("battery_source"),
            p.get("kaka_high_risk"),
            p.get("current_stock"),
            p.get("incoming"),
            p.get("pcs_sold_30d"),
            round(p.get("v_daily", 0), 4),
            p.get("multiple"),
            p.get("raw_aer"),
            p.get("qty_aer"),
            p.get("raw_tren"),
            p.get("qty_tren"),
        ])

    out_dir = os.environ.get("OUTPUT_DIR", "./output/stock_reports")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Stoc_viitor_completat_" + report_date + ".xlsx")
    wb.save(out_path)
    wb.close()

    print("=== report_writer ===")
    print("source_file :", os.path.basename(source_path))
    print("output_file :", out_path)
    print("comanda_aer scrise :", n_aer)
    print("comanda_tren scrise :", n_tren)
    print("celule Kaka marcate :", n_kaka_marked)
    print("Restock_calcul rânduri date :", len(ordered))
    if warnings:
        print("WARNINGS (" + str(len(warnings)) + "):")
        for w in warnings[:20]:
            print("   -", w)
    else:
        print("warnings : 0")


if __name__ == "__main__":
    main()
